import streamlit as st
import math
import io
import pandas as pd
from datetime import datetime
from typing import Dict, Tuple, Optional, List
import warnings
warnings.filterwarnings('ignore')

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="ONEE Tech Assistant",
    page_icon="⚡",
    layout="centered",
    initial_sidebar_state="collapsed"
)

# ── Constants ──────────────────────────────────────────────────────────────────
class Constants:
    # Resistivities (Ω·mm²/m)
    RESISTIVITY = {
        "Cuivre (Cu)": 0.0225,
        "Aluminium (Al)": 0.036
    }
    
    # Standard cable sections (mm²)
    CABLE_SECTIONS = [16, 25, 35, 50, 70, 95, 120, 150, 185, 240, 300]
    
    # Network voltages
    NETWORK_VOLTAGES = {
        "Triphasé BT (400V)": 400.0,
        "Monophasé (230V)": 230.0,
        "MT (5500V)": 5500.0
    }
    
    # Thresholds
    CHARGE_WARNING = 80.0
    CHARGE_CRITICAL = 100.0
    VOLTAGE_DROP_WARNING = 3.0
    VOLTAGE_DROP_CRITICAL = 5.0

# ── Session state init ─────────────────────────────────────────────────────────
def init_session_state():
    """Initialize session state variables"""
    defaults = {
        "res_t1": None,
        "res_t2": None,
        "calculation_history": []
    }
    for key, default in defaults.items():
        if key not in st.session_state:
            st.session_state[key] = default

init_session_state()

# ── Custom CSS ─────────────────────────────────────────────────────────────────
def load_css():
    """Load custom CSS styling"""
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow:wght@400;600;700;800&family=Barlow+Condensed:wght@700;800&display=swap');
    
    :root {
        --onee-green:  #00793B;
        --onee-dark:   #00501F;
        --onee-light:  #E8F5EE;
        --onee-accent: #F4A800;
        --onee-red:    #D32F2F;
        --bg:          #F4F6F0;
        --card:        #FFFFFF;
        --text:        #1A2E1A;
        --muted:       #5A7A5A;
        --border:      #C8DCC8;
    }
    
    * { font-family: 'Barlow', sans-serif; }
    
    .stApp {
        background: var(--bg);
        background-image:
            radial-gradient(circle at 10% 20%, rgba(0,121,59,0.07) 0%, transparent 50%),
            radial-gradient(circle at 90% 80%, rgba(244,168,0,0.06) 0%, transparent 50%);
    }
    
    /* Header */
    .onee-header {
        background: linear-gradient(135deg, var(--onee-dark) 0%, var(--onee-green) 100%);
        border-radius: 20px;
        padding: 32px 36px 28px;
        margin-bottom: 28px;
        position: relative;
        overflow: hidden;
        box-shadow: 0 8px 32px rgba(0,80,31,0.25);
        transition: transform 0.3s ease;
    }
    
    .onee-header:hover {
        transform: translateY(-2px);
    }
    
    .onee-header::before {
        content: '⚡';
        position: absolute;
        right: 28px;
        top: 50%;
        transform: translateY(-50%);
        font-size: 80px;
        opacity: 0.12;
        line-height: 1;
    }
    
    .onee-header::after {
        content: '';
        position: absolute;
        bottom: 0;
        left: 0;
        right: 0;
        height: 4px;
        background: linear-gradient(90deg, var(--onee-accent), transparent);
    }
    
    .onee-logo {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 13px;
        font-weight: 700;
        letter-spacing: 4px;
        color: rgba(255,255,255,0.6);
        text-transform: uppercase;
        margin-bottom: 6px;
    }
    
    .onee-title {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 34px;
        font-weight: 800;
        color: #FFFFFF;
        line-height: 1.1;
        margin: 0;
    }
    
    .onee-subtitle {
        font-size: 14px;
        color: rgba(255,255,255,0.65);
        margin-top: 8px;
        font-weight: 400;
    }
    
    /* Cards */
    .calc-card {
        background: var(--card);
        border: 1px solid var(--border);
        border-radius: 16px;
        padding: 28px;
        margin-bottom: 20px;
        box-shadow: 0 2px 12px rgba(0,80,31,0.07);
        transition: box-shadow 0.3s ease;
    }
    
    .calc-card:hover {
        box-shadow: 0 4px 20px rgba(0,80,31,0.12);
    }
    
    .card-title {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 20px;
        font-weight: 700;
        color: var(--onee-dark);
        margin-bottom: 20px;
        padding-bottom: 12px;
        border-bottom: 2px solid var(--onee-light);
    }
    
    /* Result boxes */
    .result-box {
        border-radius: 12px;
        padding: 20px 24px;
        margin-top: 20px;
        border-left: 5px solid;
        animation: slideIn 0.3s ease;
    }
    
    @keyframes slideIn {
        from {
            opacity: 0;
            transform: translateY(-10px);
        }
        to {
            opacity: 1;
            transform: translateY(0);
        }
    }
    
    .result-ok   { background:#E8F5EE; border-color:var(--onee-green); }
    .result-warn { background:#FFF8E1; border-color:var(--onee-accent); }
    .result-err  { background:#FFEBEE; border-color:var(--onee-red);   }
    
    .result-label {
        font-size: 11px;
        font-weight: 700;
        letter-spacing: 2px;
        text-transform: uppercase;
        margin-bottom: 4px;
        color: #333333;
    }
    
    .result-value {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 40px;
        font-weight: 800;
        line-height: 1;
        margin-bottom: 6px;
    }
    
    .result-ok   .result-value { color: #00501F; }
    .result-warn .result-value { color: #5C3D00; }
    .result-err  .result-value { color: #B71C1C; }
    
    .result-msg {
        font-size: 14px;
        font-weight: 600;
        margin-top: 8px;
        color: #1A1A1A;
    }
    
    /* Metrics */
    [data-testid="metric-container"] {
        background: var(--onee-light);
        border-radius: 10px;
        padding: 12px 16px !important;
        border: 1px solid var(--border);
        transition: transform 0.2s ease;
    }
    
    [data-testid="metric-container"]:hover {
        transform: translateY(-2px);
    }
    
    /* Tabs */
    .stTabs [data-baseweb="tab-list"] {
        background: var(--onee-light);
        border-radius: 12px;
        padding: 4px;
        gap: 4px;
        border: none;
        margin-bottom: 20px;
    }
    
    .stTabs [data-baseweb="tab"] {
        border-radius: 8px;
        font-weight: 700;
        font-size: 14px;
        color: var(--muted);
        padding: 10px 20px;
        border: none;
        background: transparent;
        transition: all 0.2s ease;
    }
    
    .stTabs [aria-selected="true"] {
        background: var(--onee-green) !important;
        color: white !important;
    }
    
    /* Buttons */
    .stButton > button {
        background: linear-gradient(135deg, var(--onee-green), var(--onee-dark)) !important;
        color: white !important;
        border: none !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 15px !important;
        padding: 12px 32px !important;
        width: 100%;
        box-shadow: 0 4px 14px rgba(0,121,59,0.3) !important;
        transition: all 0.2s ease !important;
    }
    
    .stButton > button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 20px rgba(0,121,59,0.4) !important;
    }
    
    .stButton > button:active {
        transform: translateY(0) !important;
    }
    
    /* Download buttons */
    .dl-excel .stDownloadButton > button,
    .dl-pdf .stDownloadButton > button {
        background: white !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 14px !important;
        padding: 10px 24px !important;
        width: 100%;
        transition: all 0.2s ease !important;
    }
    
    .dl-excel .stDownloadButton > button {
        color: var(--onee-green) !important;
        border: 2px solid var(--onee-green) !important;
    }
    
    .dl-pdf .stDownloadButton > button {
        color: #C62828 !important;
        border: 2px solid #C62828 !important;
    }
    
    .dl-excel .stDownloadButton > button:hover,
    .dl-pdf .stDownloadButton > button:hover {
        transform: translateY(-2px);
    }
    
    /* Footer */
    .onee-footer {
        text-align: center;
        padding: 20px;
        color: var(--muted);
        font-size: 12px;
        margin-top: 32px;
        border-top: 1px solid var(--border);
    }
    
    .badge {
        display: inline-block;
        background: var(--onee-light);
        color: var(--onee-dark);
        font-size: 11px;
        font-weight: 700;
        padding: 3px 10px;
        border-radius: 20px;
        border: 1px solid var(--border);
    }
    
    /* Hide default elements */
    #MainMenu, footer, header { visibility: hidden; }
    .block-container { padding-top: 2rem; }
    
    /* Loading spinner */
    .stSpinner {
        text-align: center;
        margin: 20px 0;
    }
    
    /* Responsive design */
    @media (max-width: 768px) {
        .onee-title {
            font-size: 24px;
        }
        
        .result-value {
            font-size: 28px;
        }
        
        .calc-card {
            padding: 20px;
        }
    }
    </style>
    """, unsafe_allow_html=True)

# ── Helper functions ───────────────────────────────────────────────────────────
def calculate_transformer_load(s_nom: float, p_reel: float, cos_phi: float) -> Dict:
    """Calculate transformer load and related parameters"""
    s_reel = p_reel / cos_phi if cos_phi > 0 else 0
    charge = (s_reel / s_nom) * 100 if s_nom > 0 else 0
    q_reel = p_reel * math.tan(math.acos(cos_phi))
    
    # Determine status
    if charge > Constants.CHARGE_CRITICAL:
        status = "critical"
        message = "⚠️ SURCHARGE — Remplacer ou décharger le transformateur immédiatement !"
    elif charge > Constants.CHARGE_WARNING:
        status = "warning"
        message = "⚡ Charge élevée — Surveillance recommandée. Prévoir renforcement."
    else:
        status = "normal"
        message = "✓ Charge normale — Transformateur dans les limites opérationnelles."
    
    return {
        "s_nom": s_nom,
        "p_reel": p_reel,
        "cos_phi": cos_phi,
        "s_reel": s_reel,
        "charge": charge,
        "q_reel": q_reel,
        "status": status,
        "message": message
    }

def calculate_voltage_drop(i_ph: Tuple[float, float, float], L: float, section: float,
                           material: str, tension_ref: float, cos_phi: float) -> Dict:
    """Calculate voltage drop and related parameters"""
    i = sum(i_ph) / 3
    rho = Constants.RESISTIVITY[material]
    R = (rho * L) / section
    delta_u = math.sqrt(3) * i * R
    perc_drop = (delta_u / tension_ref) * 100
    u_arrive = tension_ref - delta_u
    
    # Determine status and recommendations
    if perc_drop > Constants.VOLTAGE_DROP_CRITICAL:
        status = "critical"
        message = "❌ Chute > 5% — Non conforme NFC 11-201. Augmenter la section immédiatement !"
        # Calculate recommended section
        s_min = (rho * L * math.sqrt(3) * i) / (0.05 * tension_ref)
        recommended_section = next((s for s in Constants.CABLE_SECTIONS if s >= s_min), 
                                   Constants.CABLE_SECTIONS[-1])
    elif perc_drop > Constants.VOLTAGE_DROP_WARNING:
        status = "warning"
        message = "⚠️ Chute entre 3% et 5% — Acceptable mais limite. Vérifier les protections."
        recommended_section = None
    else:
        status = "normal"
        message = "✓ Chute ≤ 3% — Conforme aux normes ONEE. Installation correcte."
        recommended_section = None
    
    return {
        "i_ph": i_ph,
        "i": i,
        "L": L,
        "section": section,
        "material": material,
        "cos_phi": cos_phi,
        "tension_ref": tension_ref,
        "R": R,
        "delta_u": delta_u,
        "perc_drop": perc_drop,
        "u_arrive": u_arrive,
        "status": status,
        "message": message,
        "recommended_section": recommended_section
    }

def generate_excel_report(title: str, data_rows: List[Tuple[str, float, str]]) -> io.BytesIO:
    """Generate Excel report with proper formatting"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:30]
        
        # Define styles
        header_fill = PatternFill("solid", fgColor="00793B")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill("solid", fgColor="E8F5EE")
        section_font = Font(bold=True, color="00501F")
        border = Border(
            left=Side(style="thin", color="C8DCC8"),
            right=Side(style="thin", color="C8DCC8"),
            top=Side(style="thin", color="C8DCC8"),
            bottom=Side(style="thin", color="C8DCC8")
        )
        
        # Title
        ws.merge_cells("A1:C1")
        ws.cell(1, 1, f"ONEE — {title}").fill = header_fill
        ws.cell(1, 1).font = Font(bold=True, color="FFFFFF", size=14)
        ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 32
        
        # Date
        ws.merge_cells("A2:C2")
        ws.cell(2, 1, f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}").font = Font(italic=True, color="5A7A5A")
        ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        for col, header in enumerate(["Paramètre", "Valeur", "Unité"], 1):
            cell = ws.cell(3, col, header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        
        # Data
        for row, (param, value, unit) in enumerate(data_rows, 4):
            # Parameter column (bold)
            cell = ws.cell(row, 1, param)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="left", vertical="center")
            cell.border = border
            
            # Value column
            cell = ws.cell(row, 2, value)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # Unit column
            cell = ws.cell(row, 3, unit)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
            
            # Alternate row background
            if row % 2 == 0:
                for col in range(1, 4):
                    ws.cell(row, col).fill = section_fill
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 32
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 12
        
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf
    except Exception as e:
        st.error(f"Erreur lors de la génération du fichier Excel: {str(e)}")
        return None

def generate_pdf_report(title: str, data_rows: List[Tuple[str, float, str]], status: str) -> io.BytesIO:
    """Generate PDF report"""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                               leftMargin=2*cm, rightMargin=2*cm,
                               topMargin=2*cm, bottomMargin=2*cm)
        
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            "Title",
            fontSize=16,
            fontName="Helvetica-Bold",
            textColor=colors.white,
            spaceAfter=12
        )
        
        sub_style = ParagraphStyle(
            "Subtitle",
            fontSize=9,
            fontName="Helvetica",
            textColor=colors.HexColor("#ccddcc")
        )
        
        story = []
        
        # Header
        header_data = [
            [Paragraph(f"ONEE — {title}", title_style),
             Paragraph(f"Généré le {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub_style)]
        ]
        header_table = Table(header_data, colWidths=[11*cm, 6*cm])
        header_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#00793B")),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 16),
            ("RIGHTPADDING", (0, 0), (-1, -1), 16),
            ("TOPPADDING", (0, 0), (-1, -1), 14),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 14),
        ]))
        story.append(header_table)
        story.append(Spacer(1, 0.5*cm))
        
        # Status
        status_style = ParagraphStyle(
            "Status",
            fontSize=11,
            fontName="Helvetica-Bold",
            textColor=colors.HexColor("#00501F"),
            spaceAfter=8
        )
        story.append(Paragraph(f"Statut : {status}", status_style))
        
        # Data table
        table_data = [["Paramètre", "Valeur", "Unité"]]
        for param, value, unit in data_rows:
            table_data.append([param, str(value), unit])
        
        data_table = Table(table_data, colWidths=[8*cm, 5*cm, 4*cm])
        data_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00501F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#E8F5EE")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#C8DCC8")),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
        ]))
        story.append(data_table)
        story.append(Spacer(1, 0.8*cm))
        
        # Footer
        footer_style = ParagraphStyle(
            "Footer",
            fontSize=8,
            fontName="Helvetica",
            textColor=colors.HexColor("#5A7A5A")
        )
        story.append(Paragraph("ONEE Tech Assistant v2.1 — Outil interne BT/MT", footer_style))
        
        doc.build(story)
        buf.seek(0)
        return buf
    except Exception as e:
        st.error(f"Erreur lors de la génération du fichier PDF: {str(e)}")
        return None

def format_result_display(result: Dict, title: str) -> None:
    """Display formatted results"""
    status_map = {
        "normal": ("result-ok", "✅"),
        "warning": ("result-warn", "⚠️"),
        "critical": ("result-err", "❌")
    }
    box_class, icon = status_map.get(result["status"], ("result-ok", "📊"))
    
    st.markdown(f"""
    <div class="result-box {box_class}">
        <div class="result-label">{icon} {title}</div>
        <div class="result-value">{result.get('charge', result.get('perc_drop', 0)):.1f} %</div>
        <div class="result-msg">{result['message']}</div>
    </div>
    """, unsafe_allow_html=True)

def add_to_history(calculation_type: str, parameters: Dict, results: Dict) -> None:
    """Add calculation to history"""
    st.session_state.calculation_history.append({
        "timestamp": datetime.now(),
        "type": calculation_type,
        "parameters": parameters,
        "results": results
    })

# ── Main App ───────────────────────────────────────────────────────────────────
def main():
    load_css()
    
    # Header
    st.markdown("""
    <div class="onee-header">
        <div class="onee-logo">Office National de l'Electricite et de l'Eau Potable</div>
        <div class="onee-title">ONEE Tech Assistant</div>
        <div class="onee-subtitle">Outil de calculs électriques — Réseau Distribution BT/MT</div>
    </div>
    """, unsafe_allow_html=True)
    
    # Sidebar for history
    with st.sidebar:
        st.markdown("### 📋 Historique des calculs")
        if st.session_state.calculation_history:
            for calc in reversed(st.session_state.calculation_history[-5:]):
                st.markdown(f"""
                <div class="badge" style="margin-bottom: 5px;">
                    {calc['timestamp'].strftime('%H:%M')} - {calc['type']}
                </div>
                """, unsafe_allow_html=True)
            if st.button("🗑️ Effacer l'historique"):
                st.session_state.calculation_history = []
                st.rerun()
        else:
            st.info("Aucun calcul effectué")
    
    # Tabs
    tab1, tab2 = st.tabs(["🔌 Charge Transformateur", "📉 Chute de Tension"])
    
    # Tab 1: Transformer Load
    with tab1:
        with st.container():
            st.markdown('<div class="calc-card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">🔌 Calcul de Charge du Transformateur</div>', unsafe_allow_html=True)
            
            col1, col2 = st.columns(2)
            with col1:
                s_nom = st.number_input("Puissance nominale (kVA)", min_value=1.0, value=100.0, step=10.0,
                                       help="Puissance apparente nominale du transformateur")
            with col2:
                p_reel = st.number_input("Puissance active réelle (kW)", min_value=0.1, value=80.0, step=5.0,
                                        help="Puissance active mesurée")
            
            cos_phi = st.slider("Facteur de puissance cos φ", 0.50, 1.0, 0.85, 0.01, key="cos_phi_t1",
                               help="Facteur de puissance de la charge")
            
            if st.button("⚡ Calculer la charge", key="btn_t1", use_container_width=True):
                with st.spinner("Calcul en cours..."):
                    result = calculate_transformer_load(s_nom, p_reel, cos_phi)
                    st.session_state.res_t1 = result
                    add_to_history("Charge Transformateur", 
                                  {"S_nom": s_nom, "P_reel": p_reel, "cos_phi": cos_phi}, 
                                  result)
            
            if st.session_state.res_t1:
                result = st.session_state.res_t1
                format_result_display(result, "Taux de charge")
                
                st.markdown("---")
                col1, col2, col3 = st.columns(3)
                col1.metric("S apparente réelle", f"{result['s_reel']:.2f} kVA")
                col2.metric("Q réactive", f"{result['q_reel']:.2f} kVAR")
                col3.metric("cos φ", f"{result['cos_phi']:.2f}")
                
                # Data for export
                data_rows = [
                    ("Puissance nominale", result['s_nom'], "kVA"),
                    ("Puissance active réelle", result['p_reel'], "kW"),
                    ("Facteur de puissance", result['cos_phi'], "—"),
                    ("Puissance apparente réelle", round(result['s_reel'], 2), "kVA"),
                    ("Puissance réactive", round(result['q_reel'], 2), "kVAR"),
                    ("Taux de charge", round(result['charge'], 2), "%"),
                ]
                
                st.markdown("---")
                col1, col2 = st.columns(2)
                with col1:
                    excel_file = generate_excel_report("Rapport Charge Transformateur", data_rows)
                    if excel_file:
                        st.download_button(
                            "📊 Télécharger Excel",
                            data=excel_file,
                            file_name=f"ONEE_Charge_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_excel_t1",
                            use_container_width=True
                        )
                with col2:
                    pdf_file = generate_pdf_report("Rapport Charge Transformateur", data_rows, result['message'])
                    if pdf_file:
                        st.download_button(
                            "📄 Télécharger PDF",
                            data=pdf_file,
                            file_name=f"ONEE_Charge_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            key="dl_pdf_t1",
                            use_container_width=True
                        )
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Tab 2: Voltage Drop
    with tab2:
        with st.container():
            st.markdown('<div class="calc-card">', unsafe_allow_html=True)
            st.markdown('<div class="card-title">📉 Calcul de Chute de Tension — Triphasé</div>', unsafe_allow_html=True)
            
            st.markdown("**Courants par phase (A)**")
            col1, col2, col3 = st.columns(3)
            with col1:
                i_ph1 = st.number_input("Phase 1 (A)", min_value=0.0, value=50.0, step=1.0, key="ph1")
            with col2:
                i_ph2 = st.number_input("Phase 2 (A)", min_value=0.0, value=50.0, step=1.0, key="ph2")
            with col3:
                i_ph3 = st.number_input("Phase 3 (A)", min_value=0.0, value=50.0, step=1.0, key="ph3")
            
            i_mean = (i_ph1 + i_ph2 + i_ph3) / 3
            st.info(f"⚡ Courant moyen utilisé pour le calcul : **{i_mean:.2f} A**")
            
            L = st.number_input("Longueur du câble (m)", min_value=1.0, value=100.0, step=10.0,
                               help="Distance entre le transformateur et la charge")
            
            col1, col2 = st.columns(2)
            with col1:
                section = st.selectbox("Section du câble (mm²)", Constants.CABLE_SECTIONS,
                                       help="Section du conducteur")
            with col2:
                material = st.radio("Matériau conducteur", list(Constants.RESISTIVITY.keys()))
            
            col1, col2 = st.columns(2)
            with col1:
                network_type = st.selectbox("Type de réseau", ["Personnalisé"] + list(Constants.NETWORK_VOLTAGES.keys()))
                if network_type != "Personnalisé":
                    tension_ref = Constants.NETWORK_VOLTAGES[network_type]
                else:
                    tension_ref = st.number_input("Tension du réseau (V)", min_value=100.0, value=400.0, step=10.0)
            
            with col2:
                cos_phi = st.slider("Facteur de puissance cos φ", 0.50, 1.0, 0.85, 0.01, key="cos_phi_t2")
            
            if st.button("⚡ Calculer la chute de tension", key="btn_t2", use_container_width=True):
                with st.spinner("Calcul en cours..."):
                    result = calculate_voltage_drop(
                        (i_ph1, i_ph2, i_ph3), L, section, material, tension_ref, cos_phi
                    )
                    st.session_state.res_t2 = result
                    add_to_history("Chute de Tension",
                                  {"Courants": (i_ph1, i_ph2, i_ph3), "L": L, "Section": section},
                                  result)
            
            if st.session_state.res_t2:
                result = st.session_state.res_t2
                format_result_display(result, "Chute de tension")
                
                st.markdown("---")
                col1, col2, col3 = st.columns(3)
                col1.metric("Tension départ", f"{result['tension_ref']:.0f} V")
                col2.metric("ΔU calculé", f"{result['delta_u']:.2f} V")
                col3.metric("Tension arrivée", f"{result['u_arrive']:.1f} V")
                
                if result['recommended_section']:
                    st.info(f"💡 Section minimale recommandée : **{result['recommended_section']} mm²** (pour ΔU ≤ 5%)")
                    st.warning("⚠️ La section actuelle est insuffisante. Augmentez la section du câble.")
                
                # Data for export
                data_rows = [
                    ("Tension réseau", result['tension_ref'], "V"),
                    ("Courant Phase 1", result['i_ph'][0], "A"),
                    ("Courant Phase 2", result['i_ph'][1], "A"),
                    ("Courant Phase 3", result['i_ph'][2], "A"),
                    ("Courant moyen", round(result['i'], 2), "A"),
                    ("Longueur câble", result['L'], "m"),
                    ("Section câble", result['section'], "mm²"),
                    ("Matériau", result['material'], "—"),
                    ("cos φ", result['cos_phi'], "—"),
                    ("Résistance R", round(result['R'], 4), "Ω"),
                    ("Chute de tension dU", round(result['delta_u'], 2), "V"),
                    ("Chute en %", round(result['perc_drop'], 2), "%"),
                    ("Tension arrivée", round(result['u_arrive'], 1), "V"),
                ]
                
                st.markdown("---")
                col1, col2 = st.columns(2)
                with col1:
                    excel_file = generate_excel_report("Rapport Chute de Tension", data_rows)
                    if excel_file:
                        st.download_button(
                            "📊 Télécharger Excel",
                            data=excel_file,
                            file_name=f"ONEE_ChuteTension_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="dl_excel_t2",
                            use_container_width=True
                        )
                with col2:
                    pdf_file = generate_pdf_report("Rapport Chute de Tension", data_rows, result['message'])
                    if pdf_file:
                        st.download_button(
                            "📄 Télécharger PDF",
                            data=pdf_file,
                            file_name=f"ONEE_ChuteTension_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf",
                            mime="application/pdf",
                            key="dl_pdf_t2",
                            use_container_width=True
                        )
            
            st.markdown('</div>', unsafe_allow_html=True)
    
    # Footer
    st.markdown(f"""
    <div class="onee-footer">
        <span class="badge">ONEE Tech v2.1</span> &nbsp;|&nbsp;
        Outil interne de calculs électriques — Distribution BT/MT &nbsp;|&nbsp;
        {datetime.now().strftime('%Y')}
    </div>
    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()
